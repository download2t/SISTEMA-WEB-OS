from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q, Avg, Sum, Count
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from datetime import date, datetime, timedelta
from io import BytesIO
import locale
import json

# Imports para relatórios
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from core.views import has_permission
from governanca.models import Funcionarios, ControleQuartos, MotivoAusencia
from governanca.forms import FuncionariosForm, ControleQuartosForm, FiltroControleQuartosForm


# ======= VIEWS PARA FUNCIONÁRIOS =======

@login_required
@user_passes_test(has_permission, login_url='403')
def listar_funcionarios(request):
    """Lista todos os funcionários"""
    search = request.GET.get('search', '')
    funcionarios = Funcionarios.objects.all()
    
    if search:
        funcionarios = funcionarios.filter(
            Q(nome__icontains=search) | Q(cargo__icontains=search)
        )
    
    funcionarios = funcionarios.order_by('nome')
    
    context = {
        'funcionarios': funcionarios,
        'search': search,
    }
    return render(request, 'governanca/quartos/funcionarios/listar_funcionarios.html', context)


@login_required
@user_passes_test(has_permission, login_url='403')
def criar_funcionario(request):
    """Cria um novo funcionário"""
    if request.method == 'POST':
        form = FuncionariosForm(request.POST)
        if form.is_valid():
            funcionario = form.save()
            messages.success(request, f'Funcionário {funcionario.nome} criado com sucesso!')
            return redirect('listar_funcionarios')
        else:
            messages.error(request, 'Erro ao criar funcionário. Verifique os dados informados.')
    else:
        form = FuncionariosForm()
    
    context = {
        'form': form,
        'title': 'Criar Funcionário',
        'button_text': 'Criar Funcionário'
    }
    return render(request, 'governanca/quartos/funcionarios/form_funcionario.html', context)


@login_required
@user_passes_test(has_permission, login_url='403')
def editar_funcionario(request, funcionario_id):
    """Edita um funcionário existente"""
    funcionario = get_object_or_404(Funcionarios, id=funcionario_id)
    
    if request.method == 'POST':
        form = FuncionariosForm(request.POST, instance=funcionario)
        if form.is_valid():
            funcionario = form.save()
            messages.success(request, f'Funcionário {funcionario.nome} atualizado com sucesso!')
            return redirect('listar_funcionarios')
        else:
            messages.error(request, 'Erro ao atualizar funcionário. Verifique os dados informados.')
    else:
        form = FuncionariosForm(instance=funcionario)
    
    context = {
        'form': form,
        'funcionario': funcionario,
        'title': 'Editar Funcionário',
        'button_text': 'Atualizar Funcionário'
    }
    return render(request, 'governanca/quartos/funcionarios/form_funcionario.html', context)


@login_required
@user_passes_test(has_permission, login_url='403')
def excluir_funcionario(request, funcionario_id):
    """Exclui um funcionário"""
    funcionario = get_object_or_404(Funcionarios, id=funcionario_id)
    
    # Verifica se o funcionário possui registros de controle
    tem_registros = ControleQuartos.objects.filter(funcionario=funcionario).exists()
    
    if tem_registros:
        messages.error(request, f'Não é possível excluir {funcionario.nome} pois possui registros de controle de quartos.')
        return redirect('listar_funcionarios')
    
    funcionario.delete()
    messages.success(request, f'Funcionário {funcionario.nome} excluído com sucesso!')
    return redirect('listar_funcionarios')


# ======= VIEWS PARA CONTROLE DE QUARTOS =======

@login_required
@user_passes_test(has_permission, login_url='403')
def listar_controle_quartos(request):
    """Lista todos os registros de controle de quartos com filtros"""
    form = FiltroControleQuartosForm(request.GET)
    controles = ControleQuartos.objects.select_related('funcionario', 'motivo_ausencia').all()
    
    # Aplicar filtros
    if form.is_valid():
        if form.cleaned_data['data_inicio']:
            controles = controles.filter(data__gte=form.cleaned_data['data_inicio'])
        if form.cleaned_data['data_fim']:
            controles = controles.filter(data__lte=form.cleaned_data['data_fim'])
        if form.cleaned_data['funcionario']:
            controles = controles.filter(funcionario=form.cleaned_data['funcionario'])
    
    # Se não há filtros, mostrar apenas os últimos 30 dias
    if not any(request.GET.values()):
        data_limite = date.today() - timedelta(days=30)
        controles = controles.filter(data__gte=data_limite)
    
    # Garantir ordenação sempre por data mais recente primeiro, depois por nome do funcionário
    controles = controles.order_by('-data', 'funcionario__nome')
    
    # Calcular estatísticas baseadas nos controles filtrados
    if controles.exists():
        # Separar controles que afetam estatísticas
        controles_para_stats = [c for c in controles if c.afeta_estatisticas()]
        
        if controles_para_stats:
            total_quartos = sum(c.quantidade_quartos for c in controles_para_stats)
            total_realizados = sum(c.realizados_base for c in controles_para_stats)
            media_porcentagem = (total_realizados / total_quartos * 100) if total_quartos > 0 else 0
            
            stats = {
                'total_registros': controles.count(),
                'registros_para_media': len(controles_para_stats),
                'media_porcentagem': media_porcentagem,
                'total_quartos': total_quartos,
                'total_realizados': total_realizados,
            }
        else:
            stats = {
                'total_registros': controles.count(),
                'registros_para_media': 0,
                'media_porcentagem': 0,
                'total_quartos': 0,
                'total_realizados': 0,
            }
    else:
        stats = {
            'total_registros': 0,
            'registros_para_media': 0,
            'media_porcentagem': 0,
            'total_quartos': 0,
            'total_realizados': 0,
        }
    
    context = {
        'controles': controles,
        'form': form,
        'stats': stats,
    }
    return render(request, 'governanca/quartos/controle/listar_controle_quartos.html', context)


@login_required
@user_passes_test(has_permission, login_url='403')
def criar_controle_quartos(request):
    """Cria um novo registro de controle de quartos"""
    if request.method == 'POST':
        form = ControleQuartosForm(request.POST)
        if form.is_valid():
            controle = form.save()
            messages.success(request, f'Controle de quartos criado para {controle.funcionario.nome} em {controle.data.strftime("%d/%m/%Y")}!')
            return redirect('listar_controle_quartos')
        else:
            messages.error(request, 'Erro ao criar controle. Verifique os dados informados.')
    else:
        # Pré-preenche com a data de hoje
        form = ControleQuartosForm(initial={'data': date.today()})
    
    context = {
        'form': form,
        'title': 'Criar Controle de Quartos',
        'button_text': 'Criar Controle'
    }
    return render(request, 'governanca/quartos/controle/form_controle_quartos.html', context)


@login_required
@user_passes_test(has_permission, login_url='403')
def editar_controle_quartos(request, controle_id):
    """Edita um registro de controle de quartos existente"""
    controle = get_object_or_404(ControleQuartos, id=controle_id)
    
    if request.method == 'POST':
        form = ControleQuartosForm(request.POST, instance=controle)
        if form.is_valid():
            controle = form.save()
            messages.success(request, f'Controle de quartos atualizado para {controle.funcionario.nome}!')
            return redirect('listar_controle_quartos')
        else:
            messages.error(request, 'Erro ao atualizar controle. Verifique os dados informados.')
    else:
        form = ControleQuartosForm(instance=controle)
    
    context = {
        'form': form,
        'controle': controle,
        'title': 'Editar Controle de Quartos',
        'button_text': 'Atualizar Controle'
    }
    return render(request, 'governanca/quartos/controle/form_controle_quartos.html', context)


@login_required
@user_passes_test(has_permission, login_url='403')
def excluir_controle_quartos(request, controle_id):
    """Exclui um registro de controle de quartos"""
    controle = get_object_or_404(ControleQuartos, id=controle_id)
    
    funcionario_nome = controle.funcionario.nome
    data = controle.data.strftime("%d/%m/%Y")
    
    controle.delete()
    messages.success(request, f'Controle de quartos de {funcionario_nome} em {data} excluído com sucesso!')
    return redirect('listar_controle_quartos')


@login_required
@user_passes_test(has_permission, login_url='403')
def detalhar_controle_quartos(request, controle_id):
    """Exibe detalhes completos de um registro de controle"""
    controle = get_object_or_404(ControleQuartos, id=controle_id)
    
    context = {
        'controle': controle,
    }
    return render(request, 'governanca/quartos/controle/detalhar_controle_quartos.html', context)


@login_required
@user_passes_test(has_permission, login_url='403')
def dashboard_quartos(request):
    """Dashboard com estatísticas e gráficos do controle de quartos"""
    today = date.today()
    
    # Registros de hoje (ordenados por funcionário)
    controles_hoje = ControleQuartos.objects.filter(data=today).select_related('funcionario', 'motivo_ausencia').order_by('funcionario__nome')
    
    # Estatísticas gerais dos últimos 7 dias usando novo método
    semana_passada = today - timedelta(days=7)
    stats_semana = ControleQuartos.calcular_estatisticas_gerais(semana_passada, today)
    
    # Estatísticas por funcionário (últimos 30 dias) usando novo método
    mes_passado = today - timedelta(days=30)
    funcionarios = Funcionarios.objects.filter(ativo=True)
    stats_funcionarios = []
    
    for funcionario in funcionarios:
        stats = ControleQuartos.calcular_estatisticas_funcionario(funcionario, mes_passado, today)
        if stats['total_dias'] > 0:  # Só incluir funcionários com registros
            stats_funcionarios.append({
                'funcionario__nome': funcionario.nome,
                'funcionario__cargo': funcionario.cargo,
                'total_registros': stats['total_dias'],
                'dias_trabalho': stats['dias_trabalho'],
                'dias_ausencia_justificada': stats['dias_ausencia_justificada'],
                'dias_falta_nao_justificada': stats['dias_falta_nao_justificada'],
                'media_porcentagem': round(stats['media_performance'], 1),
                'total_quartos': stats['total_quartos'],
                'total_realizados': stats['total_realizados']
            })
    
    # Ordenar por média de performance
    stats_funcionarios.sort(key=lambda x: x['media_porcentagem'], reverse=True)
    
    # Dados para gráficos - evolução individual dos funcionários
    periodo = request.GET.get('periodo', 'semanal')  # semanal ou mensal
    
    if periodo == 'mensal':
        dias_periodo = 30
        formato_data = '%d/%m'
    else:
        dias_periodo = 7
        formato_data = '%d/%m'
    
    # Dados para o eixo X (datas do período)
    labels_periodo = []
    for i in range(dias_periodo):
        data_label = today - timedelta(days=dias_periodo - 1 - i)
        labels_periodo.append(data_label.strftime(formato_data))
    
    # Buscar funcionários ativos com registros no período
    funcionarios_com_registros = []
    cores_funcionarios = [
        '#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF',
        '#FF9F40', '#28A745', '#FFC107', '#DC3545', '#17A2B8'
    ]
    
    for idx, funcionario in enumerate(funcionarios):
        registros_funcionario = ControleQuartos.objects.filter(
            funcionario=funcionario,
            data__gte=today - timedelta(days=dias_periodo)
        ).order_by('data')
        
        if registros_funcionario.exists():
            dados_funcionario = []
            
            # Criar dados para cada dia do período
            for i in range(dias_periodo):
                data_dia = today - timedelta(days=dias_periodo - 1 - i)
                
                # Buscar registro do dia
                registro_dia = registros_funcionario.filter(data=data_dia).first()
                
                if registro_dia:
                    if registro_dia.afeta_estatisticas():
                        # Trabalho normal ou falta não justificada - incluir na linha
                        performance = float(registro_dia.porcentagem)
                    else:
                        # Folga, férias, etc. - gap na linha (null)
                        performance = None
                else:
                    # Sem registro no dia - gap na linha (null)
                    performance = None
                
                dados_funcionario.append(performance)  # Só a performance, não o objeto
            
            # Só incluir funcionário se tiver pelo menos um ponto de dados válido
            pontos_validos = [d for d in dados_funcionario if d is not None]
            if pontos_validos:
                funcionarios_com_registros.append({
                    'id': funcionario.id,  # Adicionar ID do funcionário
                    'nome': funcionario.nome,
                    'data': dados_funcionario,  # Array direto para Chart.js
                    'borderColor': cores_funcionarios[idx % len(cores_funcionarios)],
                    'backgroundColor': cores_funcionarios[idx % len(cores_funcionarios)] + '20',  # Transparência
                    'pontos_validos': len(pontos_validos)
                })
    
    # Ordenar por número de pontos válidos (funcionários com mais dados primeiro)
    funcionarios_com_registros.sort(key=lambda x: x['pontos_validos'], reverse=True)
    
    # Total de funcionários ativos
    total_funcionarios_ativos = funcionarios.count()
    
    # Preparar dados para JSON (mais limpo para Chart.js)
    funcionarios_json = json.dumps(funcionarios_com_registros) if funcionarios_com_registros else json.dumps([])
    labels_json = json.dumps(labels_periodo)
    
    context = {
        'controles_hoje': controles_hoje,
        'stats_semana': stats_semana,
        'media_porcentagem_semana': round(stats_semana['media_porcentagem'], 1),
        'stats_funcionarios': stats_funcionarios,
        'total_funcionarios_ativos': total_funcionarios_ativos,
        'funcionarios_graficos': funcionarios_com_registros,
        'funcionarios_json': funcionarios_json,
        'labels_periodo': labels_periodo,
        'labels_json': labels_json,
        'periodo_atual': periodo,
        'today': today,
    }
    return render(request, 'governanca/quartos/dashboard_quartos.html', context)


# ======= RELATÓRIOS =======

@login_required
@user_passes_test(has_permission, login_url='403')
def relatorio_controle_quartos_pdf(request):
    """Gera relatório em PDF dos controles de quartos"""
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.lib.colors import Color
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib.units import cm
    
    # Aplicar os mesmos filtros da listagem
    form = FiltroControleQuartosForm(request.GET)
    controles = ControleQuartos.objects.select_related('funcionario', 'motivo_ausencia').all()
    
    # Aplicar filtros
    filtros_aplicados = []
    if form.is_valid():
        if form.cleaned_data['data_inicio']:
            controles = controles.filter(data__gte=form.cleaned_data['data_inicio'])
            filtros_aplicados.append(f"Data início: {form.cleaned_data['data_inicio'].strftime('%d/%m/%Y')}")
        if form.cleaned_data['data_fim']:
            controles = controles.filter(data__lte=form.cleaned_data['data_fim'])
            filtros_aplicados.append(f"Data fim: {form.cleaned_data['data_fim'].strftime('%d/%m/%Y')}")
        if form.cleaned_data['funcionario']:
            controles = controles.filter(funcionario=form.cleaned_data['funcionario'])
            filtros_aplicados.append(f"Funcionário: {form.cleaned_data['funcionario'].nome}")
    
    # Se não há filtros, mostrar apenas os últimos 30 dias
    if not any(request.GET.values()):
        data_limite = date.today() - timedelta(days=30)
        controles = controles.filter(data__gte=data_limite)
        filtros_aplicados.append("Últimos 30 dias")
    
    controles = controles.order_by('-data', 'funcionario__nome')
    
    # Calcular estatísticas usando novo método
    if controles.exists():
        data_inicio = controles.last().data
        data_fim = controles.first().data
        stats_gerais = ControleQuartos.calcular_estatisticas_gerais(data_inicio, data_fim)
        stats = {
            'total_registros': stats_gerais['total_registros'],
            'media_porcentagem': stats_gerais['media_porcentagem'],
            'total_quartos': stats_gerais['total_quartos'],
            'total_realizados': stats_gerais['total_realizados']
        }
    else:
        stats = {
            'total_registros': 0,
            'media_porcentagem': 0,
            'total_quartos': 0,
            'total_realizados': 0
        }
    
    # Criar o PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_controle_quartos_{date.today().strftime("%d_%m_%Y")}.pdf"'
    
    # Configuração da página em landscape com margens adequadas
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=2*cm, bottomMargin=2*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
    story = []
    
    styles = getSampleStyleSheet()
    
    # Estilos personalizados modernos
    title_style = ParagraphStyle(
        'ModernTitle',
        parent=styles['Title'],
        fontSize=24,
        fontName='Helvetica-Bold',
        alignment=1,
        spaceAfter=20,
        spaceBefore=10,
        textColor=Color(0.2, 0.3, 0.5),  # Azul moderno
        leftIndent=0,
        rightIndent=0
    )
    
    subtitle_style = ParagraphStyle(
        'ModernSubtitle',
        parent=styles['Normal'],
        fontSize=14,
        fontName='Helvetica',
        alignment=1,
        spaceAfter=15,
        textColor=Color(0.4, 0.4, 0.4),  # Cinza moderno
    )
    
    normal_style = ParagraphStyle(
        'ModernNormal',
        parent=styles['Normal'],
        fontSize=11,
        fontName='Helvetica',
        spaceAfter=8,
        textColor=Color(0.2, 0.2, 0.2),
        leftIndent=10
    )
    
    section_style = ParagraphStyle(
        'ModernSection',
        parent=styles['Heading2'],
        fontSize=16,
        fontName='Helvetica-Bold',
        spaceAfter=12,
        spaceBefore=20,
        textColor=Color(0.2, 0.4, 0.6),
        leftIndent=0
    )
    
    # Cabeçalho moderno
    story.append(Paragraph("RELATÓRIO DE CONTROLE DE QUARTOS", title_style))
    story.append(Paragraph(f"Gerado em {timezone.now().strftime('%d de %B de %Y às %H:%M')}", subtitle_style))
    story.append(Spacer(1, 20))
    
    # Linha decorativa
    from reportlab.graphics.shapes import Drawing, Line
    from reportlab.graphics import renderPDF
    
    # Filtros aplicados em seção destacada
    if filtros_aplicados:
        story.append(Paragraph("FILTROS APLICADOS", section_style))
        for filtro in filtros_aplicados:
            story.append(Paragraph(f"• {filtro}", normal_style))
        story.append(Spacer(1, 15))
    
    # Estatísticas em layout moderno
    if stats['total_registros']:
        story.append(Paragraph("RESUMO ESTATÍSTICO", section_style))
        
        stats_data = [
            ['MÉTRICA', 'VALOR', 'MÉTRICA', 'VALOR'],
            ['Total de Registros', f"{stats['total_registros']}", 
             'Média de Performance', f"{stats['media_porcentagem']:.1f}%" if stats['media_porcentagem'] else "0%"],
            ['Total de Quartos', f"{stats['total_quartos'] or 0}", 
             'Total Realizados', f"{stats['total_realizados'] or 0}"]
        ]
        
        stats_table = Table(stats_data, colWidths=[4*cm, 3*cm, 4*cm, 3*cm])
        stats_table.setStyle(TableStyle([
            # Cabeçalho moderno
            ('BACKGROUND', (0, 0), (-1, 0), Color(0.2, 0.4, 0.6)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
            
            # Dados com alternância de cores
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 11),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [Color(0.95, 0.97, 1), colors.white]),
            
            # Bordas modernas
            ('GRID', (0, 0), (-1, -1), 1, Color(0.8, 0.8, 0.8)),
            ('LINEBELOW', (0, 0), (-1, 0), 2, Color(0.2, 0.4, 0.6)),
            
            # Padding
            ('TOPPADDING', (0, 1), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 12),
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 20))
    
    # Tabela de dados moderna
    if controles.exists():
        story.append(Paragraph("REGISTROS DETALHADOS", section_style))
        
        # Cabeçalhos
        headers = ['Data', 'Funcionário', 'Motivo', 'Quartos', 'Realizados', 'Entradas', 'Saídas', 'Reservas', '%', 'Performance']
        
        # Dados
        data = [headers]
        for controle in controles:
            motivo_nome = controle.motivo_ausencia.nome if controle.motivo_ausencia else 'Trabalho Normal'
            data.append([
                controle.data.strftime('%d/%m/%Y'),
                controle.funcionario.nome,
                motivo_nome,
                str(controle.quantidade_quartos),
                str(controle.realizados),
                str(controle.permanece_realizadas),
                str(controle.saidas_realizadas),
                str(controle.reservas_realizadas),
                f"{controle.porcentagem:.1f}%",
                controle.status_desempenho
            ])
        
        # Criar tabela com larguras ajustadas
        col_widths = [2.2*cm, 4*cm, 3*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 2.2*cm]
        table = Table(data, colWidths=col_widths)
        
        # Estilo moderno da tabela
        table.setStyle(TableStyle([
            # Cabeçalho moderno
            ('BACKGROUND', (0, 0), (-1, 0), Color(0.2, 0.4, 0.6)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Dados com estilo moderno
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            
            # Alternância de cores moderna
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [Color(0.97, 0.98, 1), colors.white]),
            
            # Bordas suaves
            ('GRID', (0, 0), (-1, -1), 0.5, Color(0.7, 0.7, 0.7)),
            ('LINEBELOW', (0, 0), (-1, 0), 2, Color(0.2, 0.4, 0.6)),
            
            # Destacar colunas importantes
            ('BACKGROUND', (-2, 1), (-2, -1), Color(0.95, 1, 0.95)),  # Coluna %
            ('BACKGROUND', (-1, 1), (-1, -1), Color(0.95, 0.95, 1)),  # Coluna Performance
        ]))
        
        story.append(table)
    else:
        story.append(Paragraph("NENHUM REGISTRO ENCONTRADO", section_style))
        story.append(Paragraph("Não foram encontrados registros para os filtros aplicados.", normal_style))
    
    # Rodapé moderno
    story.append(Spacer(1, 30))
    
    footer_style = ParagraphStyle(
        'ModernFooter',
        parent=styles['Normal'],
        fontSize=10,
        fontName='Helvetica-Oblique',
        alignment=1,
        textColor=Color(0.5, 0.5, 0.5),
        spaceAfter=0
    )
    
    footer_text = f"Sistema de Controle de Quartos | Relatório gerado em {timezone.now().strftime('%d de %B de %Y às %H:%M')}"
    story.append(Paragraph(footer_text, footer_style))
    
    # Gerar PDF
    doc.build(story)
    return response


@login_required
@user_passes_test(has_permission, login_url='403')
def relatorio_controle_quartos_excel(request):
    """Gera relatório em Excel dos controles de quartos"""
    # Aplicar os mesmos filtros da listagem
    form = FiltroControleQuartosForm(request.GET)
    controles = ControleQuartos.objects.select_related('funcionario', 'motivo_ausencia').all()
    
    # Aplicar filtros
    filtros_aplicados = []
    if form.is_valid():
        if form.cleaned_data['data_inicio']:
            controles = controles.filter(data__gte=form.cleaned_data['data_inicio'])
            filtros_aplicados.append(f"Data início: {form.cleaned_data['data_inicio'].strftime('%d/%m/%Y')}")
        if form.cleaned_data['data_fim']:
            controles = controles.filter(data__lte=form.cleaned_data['data_fim'])
            filtros_aplicados.append(f"Data fim: {form.cleaned_data['data_fim'].strftime('%d/%m/%Y')}")
        if form.cleaned_data['funcionario']:
            controles = controles.filter(funcionario=form.cleaned_data['funcionario'])
            filtros_aplicados.append(f"Funcionário: {form.cleaned_data['funcionario'].nome}")
    
    # Se não há filtros, mostrar apenas os últimos 30 dias
    if not any(request.GET.values()):
        data_limite = date.today() - timedelta(days=30)
        controles = controles.filter(data__gte=data_limite)
        filtros_aplicados.append("Últimos 30 dias")
    
    controles = controles.order_by('-data', 'funcionario__nome')
    
    # Calcular estatísticas
    stats = controles.aggregate(
        total_registros=Count('id'),
        media_porcentagem=Avg('porcentagem'),
        total_quartos=Sum('quantidade_quartos'),
        total_realizados=Sum('realizados')
    )
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Controle de Quartos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:K1')
    ws['A1'] = "RELATÓRIO DE CONTROLE DE QUARTOS"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Data de geração
    ws.merge_cells('A2:K2')
    ws['A2'] = f"Gerado em {date.today().strftime('%d/%m/%Y')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    current_row = 4
    
    # Filtros aplicados
    if filtros_aplicados:
        ws[f'A{current_row}'] = "Filtros Aplicados:"
        ws[f'A{current_row}'].font = Font(bold=True)
        current_row += 1
        
        for filtro in filtros_aplicados:
            ws[f'A{current_row}'] = f"• {filtro}"
            current_row += 1
        current_row += 1
    
    # Estatísticas
    ws[f'A{current_row}'] = "Resumo Estatístico:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    stats_data = [
        ['Métrica', 'Valor'],
        ['Total de Registros', stats['total_registros'] or 0],
        ['Média de Performance', f"{stats['media_porcentagem']:.1f}%" if stats['media_porcentagem'] else "0%"],
        ['Total de Quartos', stats['total_quartos'] or 0],
        ['Total Realizados', stats['total_realizados'] or 0]
    ]
    
    for row_data in stats_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == current_row:  # Header row
                cell.font = header_font
                cell.fill = header_fill
            cell.border = border
        current_row += 1
    
    current_row += 2
    
    # Dados detalhados
    if controles:
        ws[f'A{current_row}'] = "Registros Detalhados:"
        ws[f'A{current_row}'].font = Font(bold=True)
        current_row += 1
        
        # Cabeçalhos
        headers = [
            'Data', 'Funcionário', 'Motivo de Ausência', 'Meta Permanece', 'Meta Saídas', 
            'Meta Total', 'Permanece Realiz.', 'Saídas Realiz.', 'Reservas', 
            'Total Realizado', 'Performance (%)', 'Status'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Dados
        for controle in controles:
            motivo_nome = controle.motivo_ausencia.nome if controle.motivo_ausencia else 'Trabalho Normal'
            row_data = [
                controle.data.strftime('%d/%m/%Y'),
                controle.funcionario.nome,
                motivo_nome,
                controle.permanece_entrada,
                controle.saida_entrada,
                controle.meta_total,
                controle.permanece_realizadas,
                controle.saidas_realizadas,
                controle.reservas_realizadas,
                controle.realizados,
                round(controle.porcentagem, 1),
                controle.status_desempenho
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                if col in [4, 5, 6, 7, 8, 9, 10, 11]:  # Colunas numéricas
                    cell.alignment = Alignment(horizontal='center')
                # Destacar motivos de ausência
                elif col == 3 and motivo_nome != 'Trabalho Normal':  # Coluna do motivo
                    cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            current_row += 1
        
        # Ajustar largura das colunas
        from openpyxl.utils import get_column_letter
        column_widths = [12, 25, 20, 12, 12, 12, 15, 15, 10, 15, 15, 20]
        for col, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width
    
    # Preparar resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="relatorio_controle_quartos_{date.today().strftime("%d_%m_%Y")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
@user_passes_test(has_permission, login_url='403')
def relatorio_controle_quartos_imprimir(request):
    """Página otimizada para impressão dos controles de quartos"""
    # Aplicar os mesmos filtros da listagem
    form = FiltroControleQuartosForm(request.GET)
    controles = ControleQuartos.objects.select_related('funcionario', 'motivo_ausencia').all()
    
    # Aplicar filtros
    filtros_aplicados = []
    if form.is_valid():
        if form.cleaned_data['data_inicio']:
            controles = controles.filter(data__gte=form.cleaned_data['data_inicio'])
            filtros_aplicados.append(f"Data início: {form.cleaned_data['data_inicio'].strftime('%d/%m/%Y')}")
        if form.cleaned_data['data_fim']:
            controles = controles.filter(data__lte=form.cleaned_data['data_fim'])
            filtros_aplicados.append(f"Data fim: {form.cleaned_data['data_fim'].strftime('%d/%m/%Y')}")
        if form.cleaned_data['funcionario']:
            controles = controles.filter(funcionario=form.cleaned_data['funcionario'])
            filtros_aplicados.append(f"Funcionário: {form.cleaned_data['funcionario'].nome}")
    
    # Se não há filtros, mostrar apenas os últimos 30 dias
    if not any(request.GET.values()):
        data_limite = date.today() - timedelta(days=30)
        controles = controles.filter(data__gte=data_limite)
        filtros_aplicados.append("Últimos 30 dias")
    
    controles = controles.order_by('-data', 'funcionario__nome')
    
    # Calcular estatísticas
    stats = controles.aggregate(
        total_registros=Count('id'),
        media_porcentagem=Avg('porcentagem'),
        total_quartos=Sum('quantidade_quartos'),
        total_realizados=Sum('realizados')
    )
    
    context = {
        'controles': controles,
        'stats': stats,
        'filtros_aplicados': filtros_aplicados,
        'data_relatorio': date.today(),
    }
    return render(request, 'governanca/quartos/controle/relatorio_imprimir.html', context)


# ======= VIEWS PARA MOTIVOS DE AUSÊNCIA =======

@login_required
@user_passes_test(has_permission, login_url='403')
def listar_motivos_ausencia(request):
    """Lista todos os motivos de ausência"""
    search = request.GET.get('search', '')
    motivos = MotivoAusencia.objects.all()
    
    if search:
        motivos = motivos.filter(nome__icontains=search)
    
    motivos = motivos.order_by('nome')
    
    context = {
        'motivos': motivos,
        'search': search,
    }
    return render(request, 'governanca/quartos/motivos/listar_motivos.html', context)


@login_required
@user_passes_test(has_permission, login_url='403')
def criar_motivo_ausencia(request):
    """Cria um novo motivo de ausência"""
    if request.method == 'POST':
        nome = request.POST.get('nome')
        descricao = request.POST.get('descricao', '')
        cor = request.POST.get('cor', '#dc3545')
        
        if nome:
            MotivoAusencia.objects.create(
                nome=nome,
                descricao=descricao,
                cor=cor
            )
            messages.success(request, f'Motivo "{nome}" criado com sucesso!')
            return redirect('listar_motivos_ausencia')
        else:
            messages.error(request, 'Nome do motivo é obrigatório!')
    
    context = {
        'action': 'criar',
    }
    return render(request, 'governanca/quartos/motivos/form_motivo.html', context)


@login_required
@user_passes_test(has_permission, login_url='403')
def editar_motivo_ausencia(request, motivo_id):
    """Edita um motivo de ausência"""
    motivo = get_object_or_404(MotivoAusencia, id=motivo_id)
    
    if request.method == 'POST':
        motivo.nome = request.POST.get('nome', motivo.nome)
        motivo.descricao = request.POST.get('descricao', motivo.descricao)
        motivo.cor = request.POST.get('cor', motivo.cor)
        motivo.ativo = request.POST.get('ativo') == 'on'
        
        motivo.save()
        messages.success(request, f'Motivo "{motivo.nome}" atualizado com sucesso!')
        return redirect('listar_motivos_ausencia')
    
    context = {
        'motivo': motivo,
        'action': 'editar',
    }
    return render(request, 'governanca/quartos/motivos/form_motivo.html', context)


@login_required
@user_passes_test(has_permission, login_url='403')
def excluir_motivo_ausencia(request, motivo_id):
    """Exclui um motivo de ausência"""
    motivo = get_object_or_404(MotivoAusencia, id=motivo_id)
    
    # Verificar se é um motivo do sistema
    if motivo.sistema:
        messages.error(request, f'O motivo "{motivo.nome}" é um motivo do sistema e não pode ser excluído.')
        return redirect('listar_motivos_ausencia')
    
    # Verificar se o motivo está sendo usado
    controles_usando = ControleQuartos.objects.filter(motivo_ausencia=motivo).count()
    
    if controles_usando > 0:
        messages.error(request, f'Não é possível excluir o motivo "{motivo.nome}" pois está sendo usado em {controles_usando} controle(s).')
        return redirect('listar_motivos_ausencia')
    
    nome_motivo = motivo.nome
    motivo.delete()
    messages.success(request, f'Motivo "{nome_motivo}" excluído com sucesso!')
    return redirect('listar_motivos_ausencia')


@login_required
@user_passes_test(has_permission, login_url='403')
def exportar_grafico_performance_pdf(request):
    """Exporta dados de performance em PDF com tabela detalhada e resumo estatístico"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.colors import Color, black, white
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    today = date.today()
    
    # Parâmetros da requisição
    periodo = request.GET.get('periodo', 'semanal')
    funcionarios_selecionados = request.GET.getlist('funcionarios[]')  # IDs dos funcionários selecionados
    
    if periodo == 'mensal':
        dias_periodo = 30
        formato_data = '%d/%m'
        titulo_periodo = "Últimos 30 Dias"
    else:
        dias_periodo = 7
        formato_data = '%d/%m'
        titulo_periodo = "Últimos 7 Dias"
    
    # Dados para o eixo X (datas do período)
    labels_periodo = []
    datas_periodo = []
    for i in range(dias_periodo):
        data_label = today - timedelta(days=dias_periodo - 1 - i)
        labels_periodo.append(data_label.strftime(formato_data))
        datas_periodo.append(data_label)
    
    # Buscar funcionários ativos
    funcionarios = Funcionarios.objects.filter(ativo=True)
    if funcionarios_selecionados:
        funcionarios = funcionarios.filter(id__in=funcionarios_selecionados)
    
    # Cores para os funcionários
    cores_funcionarios = [
        '#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF',
        '#FF9F40', '#28A745', '#FFC107', '#DC3545', '#17A2B8'
    ]
    
    # Preparar dados do gráfico
    funcionarios_dados = []
    for idx, funcionario in enumerate(funcionarios):
        registros_funcionario = ControleQuartos.objects.filter(
            funcionario=funcionario,
            data__gte=today - timedelta(days=dias_periodo)
        ).order_by('data')
        
        if registros_funcionario.exists():
            dados_performance = []
            registros_detalhados = []
            
            for data_dia in datas_periodo:
                registro_dia = registros_funcionario.filter(data=data_dia).first()
                
                if registro_dia and registro_dia.afeta_estatisticas():
                    performance = float(registro_dia.porcentagem)
                    registros_detalhados.append({
                        'data': data_dia.strftime(formato_data),
                        'performance': performance,
                        'quartos': registro_dia.quantidade_quartos,
                        'realizados': registro_dia.realizados,
                        'motivo': registro_dia.motivo_ausencia.nome if registro_dia.motivo_ausencia else 'Trabalho Normal'
                    })
                else:
                    performance = None
                
                dados_performance.append(performance)
            
            # Só incluir funcionário se tiver dados válidos
            pontos_validos = [d for d in dados_performance if d is not None]
            if pontos_validos:
                funcionarios_dados.append({
                    'nome': funcionario.nome,
                    'cargo': funcionario.cargo,
                    'dados': dados_performance,
                    'registros_detalhados': registros_detalhados,
                    'cor': cores_funcionarios[idx % len(cores_funcionarios)],
                    'pontos_validos': len(pontos_validos),
                    'media': sum(pontos_validos) / len(pontos_validos) if pontos_validos else 0,
                    'melhor_dia': max(pontos_validos) if pontos_validos else 0,
                    'pior_dia': min(pontos_validos) if pontos_validos else 0
                })
    
    if not funcionarios_dados:
        # Retornar erro se não há dados
        return HttpResponse("Nenhum dado encontrado para o período selecionado.", status=400)
    
    # Criar PDF com margens otimizadas
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_performance_{periodo}_{today.strftime("%d_%m_%Y")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=1.2*cm, bottomMargin=1.2*cm, leftMargin=0.8*cm, rightMargin=0.8*cm)
    story = []
    
    styles = getSampleStyleSheet()
    
    # Estilos modernos para o dashboard
    title_style = ParagraphStyle(
        'DashboardTitle',
        parent=styles['Title'],
        fontSize=26,
        fontName='Helvetica-Bold',
        alignment=1,
        spaceAfter=25,
        spaceBefore=10,
        textColor=Color(0.1, 0.3, 0.6),
        leftIndent=0,
        rightIndent=0
    )
    
    subtitle_style = ParagraphStyle(
        'DashboardSubtitle',
        parent=styles['Normal'],
        fontSize=16,
        fontName='Helvetica',
        alignment=1,
        spaceAfter=20,
        textColor=Color(0.4, 0.4, 0.4),
    )
    
    section_style = ParagraphStyle(
        'DashboardSection',
        parent=styles['Heading2'],
        fontSize=18,
        fontName='Helvetica-Bold',
        spaceAfter=15,
        spaceBefore=25,
        textColor=Color(0.1, 0.4, 0.7),
        leftIndent=0
    )
    
    normal_style = ParagraphStyle(
        'DashboardNormal',
        parent=styles['Normal'],
        fontSize=12,
        fontName='Helvetica',
        spaceAfter=10,
        textColor=Color(0.2, 0.2, 0.2),
        leftIndent=15
    )
    
    # Título moderno do dashboard
    story.append(Paragraph(f"RELATÓRIO DE PERFORMANCE - {titulo_periodo.upper()}", title_style))
    story.append(Paragraph(f"Gerado em {today.strftime('%d de %B de %Y às %H:%M')}", subtitle_style))
    story.append(Spacer(1, 20))
    
    # Informações do período com destaque
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=13,
        fontName='Helvetica',
        spaceAfter=8,
        textColor=Color(0.3, 0.3, 0.3),
        leftIndent=10
    )
    
    story.append(Paragraph(f"📅 <b>Período:</b> {datas_periodo[0].strftime('%d/%m/%Y')} a {datas_periodo[-1].strftime('%d/%m/%Y')}", info_style))
    story.append(Paragraph(f"👥 <b>Funcionários incluídos:</b> {len(funcionarios_dados)}", info_style))
    story.append(Spacer(1, 25))
    
    # Resumo estatístico moderno
    story.append(Paragraph("📊 RESUMO ESTATÍSTICO EXECUTIVO", section_style))
    story.append(Spacer(1, 10))
    
    # Tabela de estatísticas com design moderno - CORRIGIDA
    stats_data = [['FUNCIONÁRIO', 'CARGO', 'DIAS VÁLIDOS', 'MÉDIA (%)', 'CLASSIFICAÇÃO']]
    
    for func_data in sorted(funcionarios_dados, key=lambda x: x['media'], reverse=True):
        if func_data['media'] >= 90:
            status = '⭐ Excelente'
        elif func_data['media'] >= 80:
            status = '✅ Boa'
        elif func_data['media'] >= 70:
            status = '⚠️ Regular'
        else:
            status = '🔴 Baixa'
        
        stats_data.append([
            func_data['nome'],
            func_data['cargo'],
            str(func_data['pontos_validos']),
            f"{func_data['media']:.1f}%",
            status
        ])
    
    # Larguras das colunas otimizadas para aproveitar melhor o espaço (cerca de 27.7cm úteis)
    stats_table = Table(stats_data, colWidths=[8*cm, 6*cm, 3.5*cm, 3.5*cm, 4.5*cm])
    stats_table.setStyle(TableStyle([
        # Cabeçalho moderno
        ('BACKGROUND', (0, 0), (-1, 0), Color(0.1, 0.3, 0.6)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Funcionário alinhado à esquerda
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),  # Cabeçalho um pouco maior
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),  # Mais padding no cabeçalho
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Dados com cores alternadas
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [Color(0.96, 0.98, 1), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.8, Color(0.7, 0.7, 0.7)),
        ('LINEBELOW', (0, 0), (-1, 0), 2, Color(0.1, 0.3, 0.6)),
        ('TOPPADDING', (0, 1), (-1, -1), 8),  # Mais espaço nas células
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        
        # Destacar colunas importantes
        ('BACKGROUND', (3, 1), (3, -1), Color(0.95, 1, 0.95)),  # Média
        ('BACKGROUND', (-1, 1), (-1, -1), Color(0.95, 0.95, 1)),  # Classificação
        ('FONTNAME', (3, 1), (3, -1), 'Helvetica-Bold'),  # Média em negrito
    ]))
    
    story.append(stats_table)
    story.append(Spacer(1, 25))
    
    # Dados detalhados por período
    story.append(Paragraph("📈 DADOS DETALHADOS POR PERÍODO", section_style))
    story.append(Spacer(1, 10))
    
    # Criar tabela de performance por dia com design moderno
    header_row = ['FUNCIONÁRIO'] + labels_periodo + ['MÉDIA']
    detailed_data = [header_row]
    
    for func_data in sorted(funcionarios_dados, key=lambda x: x['media'], reverse=True):
        row = [func_data['nome']]
        
        for performance in func_data['dados']:
            if performance is not None:
                row.append(f"{performance:.1f}%")
            else:
                row.append("-")  # Gap (folga/ausência justificada)
        
        row.append(f"{func_data['media']:.1f}%")
        detailed_data.append(row)
    
    # Calcular larguras das colunas dinamicamente baseadas no período e espaço disponível
    # Largura total disponível: cerca de 27.7cm (A4 landscape com margens reduzidas)
    largura_total_disponivel = 27.7*cm
    
    if dias_periodo == 7:
        # Para 7 dias: mais espaço para cada coluna
        largura_funcionario = 6*cm
        largura_dados = (largura_total_disponivel - largura_funcionario - 3*cm) / dias_periodo  # 3cm para média
        largura_media = 3*cm
        col_widths = [largura_funcionario] + [largura_dados] * dias_periodo + [largura_media]
    else:
        # Para 30 dias: otimizar espaço mantendo legibilidade
        largura_funcionario = 4.5*cm
        largura_media = 2.5*cm
        largura_dados = (largura_total_disponivel - largura_funcionario - largura_media) / dias_periodo
        # Mínimo de 0.6cm por coluna de dados
        if largura_dados < 0.6*cm:
            largura_dados = 0.6*cm
        col_widths = [largura_funcionario] + [largura_dados] * dias_periodo + [largura_media]
    
    detailed_table = Table(detailed_data, colWidths=col_widths)
    detailed_table.setStyle(TableStyle([
        # Cabeçalho moderno
        ('BACKGROUND', (0, 0), (-1, 0), Color(0.2, 0.2, 0.2)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Nomes dos funcionários à esquerda
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8 if dias_periodo > 15 else 9),  # Fonte adaptativa
        ('FONTSIZE', (0, 1), (-1, -1), 7 if dias_periodo > 15 else 8),  # Fonte menor para muitos dados
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        
        # Dados com estilo moderno e compacto
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, Color(0.97, 0.97, 0.97)]),
        ('GRID', (0, 0), (-1, -1), 0.5, Color(0.6, 0.6, 0.6)),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 3),  # Padding muito reduzido
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),  # Menos padding horizontal
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        
        # Destacar coluna de funcionário e média
        ('BACKGROUND', (0, 1), (0, -1), Color(0.95, 0.95, 0.95)),  # Coluna funcionário
        ('BACKGROUND', (-1, 1), (-1, -1), Color(0.9, 0.95, 1.0)),  # Coluna média
        ('FONTNAME', (-1, 1), (-1, -1), 'Helvetica-Bold'),  # Média em negrito
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),  # Funcionários em negrito
        ('LINEBELOW', (0, 0), (-1, 0), 2, Color(0.2, 0.2, 0.2)),
    ]))
    
    story.append(detailed_table)
    story.append(Spacer(1, 30))
    
    # ADICIONAR INFORMAÇÃO VISUAL EM LUGAR DO GRÁFICO
    story.append(Paragraph("📊 ANÁLISE VISUAL DA PERFORMANCE", section_style))
    story.append(Spacer(1, 15))
    
    # Criar uma representação textual visual dos dados
    visual_data = []
    visual_data.append(['FUNCIONÁRIO', 'GRÁFICO DE BARRAS VISUAL', 'MÉDIA'])
    
    # Funcionários para visualização (top 5)
    funcionarios_para_visual = sorted(funcionarios_dados, key=lambda x: x['media'], reverse=True)[:5]
    
    for func_data in funcionarios_para_visual:
        # Criar barra visual usando caracteres
        performance = func_data['media']
        barra_tamanho = int(performance / 5)  # 100% = 20 caracteres
        barra_visual = '█' * barra_tamanho + '░' * (20 - barra_tamanho)
        
        # Adicionar cor via cor hex para identificação
        linha_visual = [
            func_data['nome'],
            f"{barra_visual} ({performance:.1f}%)",
            f"{performance:.1f}%"
        ]
        visual_data.append(linha_visual)
    
    if visual_data:
        visual_table = Table(visual_data, colWidths=[5*cm, 8*cm, 3*cm])
        visual_table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), Color(0.1, 0.3, 0.6)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Dados
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTNAME', (1, 1), (1, -1), 'Courier'),  # Fonte monospace para barras
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [Color(0.96, 0.98, 1), colors.white]),
            ('GRID', (0, 0), (-1, -1), 0.5, Color(0.7, 0.7, 0.7)),
            ('LINEBELOW', (0, 0), (-1, 0), 2, Color(0.1, 0.3, 0.6)),
            ('TOPPADDING', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            
            # Destacar média
            ('BACKGROUND', (-1, 1), (-1, -1), Color(0.95, 1, 0.95)),
            ('FONTNAME', (-1, 1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        story.append(visual_table)
        
        # Adicionar legenda das barras
        story.append(Spacer(1, 15))
        legenda_barras = ParagraphStyle(
            'LegendaBarras',
            parent=styles['Normal'],
            fontSize=10,
            fontName='Helvetica',
            textColor=Color(0.3, 0.3, 0.3),
            leftIndent=15,
            spaceAfter=8
        )
        
        story.append(Paragraph("📏 <b>Legenda das Barras Visuais:</b> █ = Desempenho alcançado | ░ = Meta não atingida | Escala: 0-100%", legenda_barras))
    else:
        story.append(Paragraph("Não há dados suficientes para gerar o gráfico.", normal_style))
    
    story.append(Spacer(1, 25))
    
    # Legenda e observações modernizadas
    story.append(Paragraph("ℹ️ LEGENDA E INFORMAÇÕES IMPORTANTES", section_style))
    
    nota_style = ParagraphStyle(
        'ModernNota',
        parent=styles['Normal'],
        fontSize=11,
        fontName='Helvetica',
        textColor=Color(0.3, 0.3, 0.3),
        leftIndent=25,
        rightIndent=25,
        spaceAfter=8,
        bulletIndent=20
    )
    
    nota_text = """
    <b>📋 Interpretação dos Dados:</b><br/>
    • <b>Dias Válidos:</b> Considera apenas dias de trabalho efetivo e faltas não justificadas<br/>
    • <b>Símbolo "-":</b> Representa folgas, férias e faltas justificadas (não impactam na média)<br/>
    • <b>Performance:</b> Percentual de quartos concluídos em relação à meta diária estabelecida<br/>
    • <b>Classificações:</b> ⭐ Excelente (≥90%), ✅ Boa (80-89%), ⚠️ Regular (70-79%), 🔴 Baixa (<70%)<br/>
    • <b>Cálculo da Média:</b> Baseado exclusivamente nos dias válidos de trabalho<br/>
    • <b>Gráfico:</b> Visualiza a evolução dos 5 funcionários com melhor performance média
    """
    story.append(Paragraph(nota_text, nota_style))
    
    # Rodapé moderno e elegante
    story.append(Spacer(1, 25))
    
    footer_style = ParagraphStyle(
        'DashboardFooter',
        parent=styles['Normal'],
        fontSize=11,
        fontName='Helvetica-Oblique',
        alignment=1,
        textColor=Color(0.4, 0.4, 0.4),
        spaceAfter=0,
        spaceBefore=10
    )
    
    footer_text = f"🏢 Sistema de Controle de Quartos | 📊 Dashboard Analytics | 📅 Gerado em {today.strftime('%d de %B de %Y às %H:%M')}"
    story.append(Paragraph(footer_text, footer_style))
    
    # Gerar PDF
    doc.build(story)
    return response


# ======= API VIEWS (AJAX) =======

@login_required
def api_controle_quartos_data(request):
    """API para retornar dados de controle de quartos em JSON"""
    if request.method == 'GET':
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')
        funcionario_id = request.GET.get('funcionario_id')
        
        controles = ControleQuartos.objects.select_related('funcionario').all()
        
        if data_inicio:
            controles = controles.filter(data__gte=data_inicio)
        if data_fim:
            controles = controles.filter(data__lte=data_fim)
        if funcionario_id:
            controles = controles.filter(funcionario_id=funcionario_id)
        
        dados = []
        for controle in controles.order_by('-data', 'funcionario__nome'):
            dados.append({
                'id': controle.id,
                'data': controle.data.strftime('%Y-%m-%d'),
                'funcionario': controle.funcionario.nome,
                'cargo': controle.funcionario.cargo,
                'porcentagem': round(controle.porcentagem, 1),
                'realizados': controle.realizados,
                'quantidade_quartos': controle.quantidade_quartos,
                'status': controle.status_desempenho
            })
        
        return JsonResponse({'dados': dados})
    
    return JsonResponse({'erro': 'Método não permitido'}, status=405)
